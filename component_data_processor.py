#!/usr/bin/env python3
"""
Component Data Processor

This script automates the processing and updating of component data based on Excel files.
It performs data cleaning, XLOOKUP processing against a Master BOM, and generates updated outputs.

Author: Auto-generated
Date: 2025-07-28
"""

import pandas as pd
import numpy as np
import logging
from datetime import datetime
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment


class ComponentDataProcessor:
    """Main class for processing component data from Excel files."""
    
    def __init__(self, config_path: str = "config.py"):
        """Initialize the processor with configuration."""
        self.setup_logging()
        self.load_config(config_path)
        self.excluded_rows = []
        self.processing_summary = {
            'total_rows': 0,
            'cleaned_rows': 0,
            'excluded_rows': 0,
            'status_d_updates': 0,
            'status_0_duplicates': 0,
            'status_nan_unknowns': 0,
            'status_x_skipped': 0
        }
    
    def setup_logging(self):
        """Setup logging configuration."""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(f'component_processor_{datetime.now().strftime("%Y%m%d")}.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def load_config(self, config_path: str):
        """Load configuration from config file."""
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("config", config_path)
            config = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(config)
            
            self.master_bom_path = config.MASTER_BOM_PATH
            self.output_dir = Path(config.OUTPUT_DIR)
            self.column_mapping = config.COLUMN_MAPPING
            self.required_columns = config.REQUIRED_COLUMNS
            
        except Exception as e:
            self.logger.warning(f"Could not load config file: {e}. Using defaults.")
            self.set_default_config()
    
    def set_default_config(self):
        """Set default configuration values."""
        self.master_bom_path = "Master_BOM.xlsx"
        self.output_dir = Path("output")
        self.column_mapping = {
            'PN': 'PN',
            'Project': 'Project',
            'Price': 'Price',
            'Supplier': 'Supplier',
            'Description': 'Description',
            'Status': 'Status'
        }
        self.required_columns = ['PN', 'Project']
    
    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean the input data according to business rules.
        
        Args:
            df: Input DataFrame
            
        Returns:
            Cleaned DataFrame
        """
        self.logger.info("Starting data cleaning process...")
        original_count = len(df)
        self.processing_summary['total_rows'] = original_count
        
        # Store original for excluded rows tracking
        df_original = df.copy()
        
        # Remove rows with missing critical values
        before_critical = len(df)
        df = df.dropna(subset=self.required_columns, how='any')
        after_critical = len(df)
        
        # Track excluded rows
        if before_critical > after_critical:
            excluded_mask = df_original[self.required_columns].isna().any(axis=1)
            self.excluded_rows.extend(df_original[excluded_mask].to_dict('records'))
        
        # Clean string columns
        string_columns = ['PN', 'Project', 'Supplier', 'Description']
        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().str.upper()
                # Remove extra spaces
                df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
                # Normalize special characters (basic ASCII conversion)
                df[col] = df[col].apply(self._normalize_text)
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        cleaned_count = len(df)
        excluded_count = original_count - cleaned_count
        
        self.processing_summary['cleaned_rows'] = cleaned_count
        self.processing_summary['excluded_rows'] = excluded_count
        
        self.logger.info(f"Data cleaning completed. Original: {original_count}, "
                        f"Cleaned: {cleaned_count}, Excluded: {excluded_count}")
        
        return df
    
    def _normalize_text(self, text: str) -> str:
        """Normalize text by removing non-ASCII characters and standardizing format."""
        if pd.isna(text) or text == 'nan':
            return ''
        
        # Convert to string and remove non-ASCII characters
        text = str(text).encode('ascii', 'ignore').decode('ascii')
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text).strip()
        return text
    
    def load_master_bom(self) -> pd.DataFrame:
        """Load the Master BOM reference file."""
        try:
            master_bom = pd.read_excel(self.master_bom_path)
            self.logger.info(f"Master BOM loaded successfully with {len(master_bom)} records")
            return master_bom
        except Exception as e:
            self.logger.error(f"Failed to load Master BOM: {e}")
            raise
    
    def perform_lookup(self, input_df: pd.DataFrame, master_bom: pd.DataFrame) -> pd.DataFrame:
        """
        Perform XLOOKUP between input data and Master BOM.
        
        Args:
            input_df: Cleaned input DataFrame
            master_bom: Master BOM DataFrame
            
        Returns:
            DataFrame with lookup results and processing actions
        """
        self.logger.info("Starting lookup process...")
        
        # Create lookup key for both dataframes
        input_df['lookup_key'] = input_df['PN'].astype(str) + '_' + input_df['Project'].astype(str)
        master_bom['lookup_key'] = master_bom['PN'].astype(str) + '_' + master_bom['Project'].astype(str)
        
        # Perform the lookup
        lookup_result = input_df.merge(
            master_bom[['lookup_key', 'Status']],
            on='lookup_key',
            how='left',
            suffixes=('', '_master')
        )
        
        # Add Notes column for processing comments
        lookup_result['Notes'] = ''
        lookup_result['Action'] = ''
        
        return lookup_result

    def process_lookup_results(self, df: pd.DataFrame, master_bom: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Process lookup results according to business logic.

        Args:
            df: DataFrame with lookup results
            master_bom: Master BOM DataFrame for updates

        Returns:
            Tuple of (processed_df, updated_master_bom)
        """
        self.logger.info("Processing lookup results...")

        additional_rows = []

        for idx, row in df.iterrows():
            status = row.get('Status', np.nan)

            if pd.isna(status):
                # Case 3: Status = "NaN" - Unknown component
                self._handle_status_nan(row, additional_rows)

            elif str(status).upper() == 'D':
                # Case 1: Status = "D" - Deprecated component
                self._handle_status_d(row, df, master_bom, idx)

            elif str(status) == '0':
                # Case 2: Status = "0" - Duplicate or uncertain match
                self._handle_status_0(row, additional_rows)

            elif str(status).upper() == 'X':
                # Case 4: Status = "X" - Already marked as old
                self._handle_status_x(row, df, idx)

        # Add additional rows to the dataframe
        if additional_rows:
            additional_df = pd.DataFrame(additional_rows)
            df = pd.concat([df, additional_df], ignore_index=True)

        self.logger.info("Lookup processing completed")
        return df, master_bom

    def _handle_status_d(self, row: pd.Series, df: pd.DataFrame, master_bom: pd.DataFrame, idx: int):
        """Handle Status = 'D' case - Update to 'X' and add comment."""
        # Update Master BOM
        lookup_key = row['lookup_key']
        master_bom.loc[master_bom['lookup_key'] == lookup_key, 'Status'] = 'X'

        # Add comment to current row
        df.at[idx, 'Notes'] = 'Status D updated to X'
        df.at[idx, 'Action'] = 'Updated'

        self.processing_summary['status_d_updates'] += 1
        self.logger.info(f"Updated status D to X for PN: {row['PN']}, Project: {row['Project']}")

    def _handle_status_0(self, row: pd.Series, additional_rows: List[Dict]):
        """Handle Status = '0' case - Add duplicate row for manual verification."""
        new_row = {
            'PN': row['PN'],
            'Project': row['Project'],
            'Price': '',
            'Description': '',
            'Supplier': '',
            'Notes': 'Duplicate or uncertain match - manual verification needed',
            'Action': 'Duplicate_Added',
            'Status': '0'
        }
        additional_rows.append(new_row)

        self.processing_summary['status_0_duplicates'] += 1
        self.logger.info(f"Added duplicate row for PN: {row['PN']}, Project: {row['Project']}")

    def _handle_status_nan(self, row: pd.Series, additional_rows: List[Dict]):
        """Handle Status = 'NaN' case - Add unknown component row."""
        new_row = {
            'PN': row['PN'],
            'Project': row['Project'],
            'Price': '',
            'Description': '',
            'Supplier': '',
            'Notes': 'Unknown PN – potential new entry',
            'Action': 'Unknown_Added',
            'Status': 'NaN'
        }
        additional_rows.append(new_row)

        self.processing_summary['status_nan_unknowns'] += 1
        self.logger.info(f"Added unknown component row for PN: {row['PN']}, Project: {row['Project']}")

    def _handle_status_x(self, row: pd.Series, df: pd.DataFrame, idx: int):
        """Handle Status = 'X' case - Skip processing."""
        df.at[idx, 'Notes'] = 'Component already marked as old - skipped'
        df.at[idx, 'Action'] = 'Skipped'

        self.processing_summary['status_x_skipped'] += 1
        self.logger.info(f"Skipped component with status X: PN: {row['PN']}, Project: {row['Project']}")

    def save_outputs(self, processed_df: pd.DataFrame, updated_master_bom: pd.DataFrame):
        """Save all output files."""
        self.logger.info("Saving output files...")

        # Create output directory if it doesn't exist
        self.output_dir.mkdir(exist_ok=True)

        # Generate timestamp for filenames
        timestamp = datetime.now().strftime("%Y-%m-%d")

        # Save updated Excel file
        output_file = self.output_dir / f"Update_{timestamp}.xlsx"
        self._save_excel_with_formatting(processed_df, output_file)

        # Save updated Master BOM
        master_bom_output = self.output_dir / f"Master_BOM_Updated_{timestamp}.xlsx"
        updated_master_bom.to_excel(master_bom_output, index=False)

        # Save excluded rows if any
        if self.excluded_rows:
            excluded_file = self.output_dir / f"Clean_Excluded_{timestamp}.xlsx"
            excluded_df = pd.DataFrame(self.excluded_rows)
            excluded_df.to_excel(excluded_file, index=False)
            self.logger.info(f"Excluded rows saved to: {excluded_file}")

        # Save summary report
        self._save_summary_report(timestamp)

        self.logger.info(f"Main output saved to: {output_file}")
        self.logger.info(f"Updated Master BOM saved to: {master_bom_output}")

    def _save_excel_with_formatting(self, df: pd.DataFrame, output_file: Path):
        """Save Excel file with formatting and comments."""
        # Remove lookup_key column before saving
        df_output = df.drop(columns=['lookup_key'], errors='ignore')

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_output.to_excel(writer, sheet_name='Updated_Data', index=False)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Updated_Data']

            # Define fill colors
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

            # Apply formatting based on Action column
            if 'Action' in df_output.columns:
                action_col_idx = df_output.columns.get_loc('Action') + 1

                for row_idx, action in enumerate(df_output['Action'], start=2):
                    if action in ['Duplicate_Added', 'Unknown_Added']:
                        # Highlight entire row in red for duplicates/unknowns
                        for col_idx in range(1, len(df_output.columns) + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
                    elif action == 'Updated':
                        # Highlight entire row in yellow for updates
                        for col_idx in range(1, len(df_output.columns) + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill

    def _save_summary_report(self, timestamp: str):
        """Save a summary report of the processing."""
        summary_file = self.output_dir / f"Processing_Summary_{timestamp}.csv"

        summary_data = {
            'Metric': list(self.processing_summary.keys()),
            'Count': list(self.processing_summary.values())
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_csv(summary_file, index=False)

        self.logger.info(f"Summary report saved to: {summary_file}")

    def process_file(self, input_file_path: str) -> bool:
        """
        Main method to process a single input file.

        Args:
            input_file_path: Path to the input Excel file

        Returns:
            True if processing was successful, False otherwise
        """
        try:
            self.logger.info(f"Starting processing of file: {input_file_path}")

            # Step 1: Load input data
            input_df = pd.read_excel(input_file_path)
            self.logger.info(f"Loaded input file with {len(input_df)} rows")

            # Step 2: Clean data
            cleaned_df = self.clean_data(input_df)

            # Step 3: Load Master BOM
            master_bom = self.load_master_bom()

            # Step 4: Perform lookup
            lookup_df = self.perform_lookup(cleaned_df, master_bom)

            # Step 5: Process results according to business logic
            processed_df, updated_master_bom = self.process_lookup_results(lookup_df, master_bom)

            # Step 6: Save outputs
            self.save_outputs(processed_df, updated_master_bom)

            # Print completion message
            print("Update completed successfully")
            self.logger.info("Processing completed successfully")

            # Print summary
            self._print_summary()

            return True

        except Exception as e:
            self.logger.error(f"Error processing file: {e}")
            print(f"Error: {e}")
            return False

    def _print_summary(self):
        """Print a summary of the processing results."""
        print("\n" + "="*50)
        print("PROCESSING SUMMARY")
        print("="*50)
        for metric, count in self.processing_summary.items():
            print(f"{metric.replace('_', ' ').title()}: {count}")
        print("="*50)


def main():
    """Main function for CLI usage."""
    import argparse

    parser = argparse.ArgumentParser(description='Component Data Processor')
    parser.add_argument('input_file', help='Path to the input Excel file')
    parser.add_argument('--config', default='config.py', help='Path to configuration file')

    args = parser.parse_args()

    # Create processor instance
    processor = ComponentDataProcessor(args.config)

    # Process the file
    success = processor.process_file(args.input_file)

    if success:
        print("\nProcessing completed successfully!")
    else:
        print("\nProcessing failed. Check the log file for details.")
        exit(1)


if __name__ == "__main__":
    main()
