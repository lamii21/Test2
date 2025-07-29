"""
Gestionnaire Excel pour le Component Data Processor.

Fournit des fonctionnalités pour lire, écrire et formater des fichiers Excel
avec support pour les commentaires, la coloration et les styles.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..utils.logger import Logger
from ..utils.file_manager import FileManager


class ExcelHandler:
    """Gestionnaire pour les opérations Excel avancées."""
    
    def __init__(self, logger: Optional[Logger] = None):
        """
        Initialise le gestionnaire Excel.
        
        Args:
            logger: Instance de logger (optionnel)
        """
        self.logger = logger or Logger("ExcelHandler")
        self.file_manager = FileManager()
        
        # Définir les couleurs de formatage
        self.colors = {
            'duplicate': PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
            'updated': PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            'skipped': PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
            'error': PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
            'header': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        }
        
        # Définir les polices
        self.fonts = {
            'header': Font(color="FFFFFF", bold=True),
            'bold': Font(bold=True),
            'normal': Font()
        }
    
    def read_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Lit un fichier Excel et retourne un DataFrame.
        
        Args:
            file_path: Chemin du fichier Excel
            sheet_name: Nom de la feuille (optionnel)
            
        Returns:
            DataFrame avec les données
            
        Raises:
            Exception: Si erreur lors de la lecture
        """
        try:
            self.logger.info(f"Lecture du fichier Excel: {file_path}")
            
            # Lire le fichier Excel
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            self.logger.info(f"Fichier lu avec succès: {len(df)} lignes, {len(df.columns)} colonnes")
            return df
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la lecture du fichier Excel: {e}")
            raise
    
    def write_excel_file(self, df: pd.DataFrame, file_path: str, sheet_name: str = "Data") -> bool:
        """
        Écrit un DataFrame dans un fichier Excel simple.
        
        Args:
            df: DataFrame à écrire
            file_path: Chemin du fichier de sortie
            sheet_name: Nom de la feuille
            
        Returns:
            True si succès, False sinon
        """
        try:
            self.logger.info(f"Écriture du fichier Excel: {file_path}")
            
            df.to_excel(file_path, sheet_name=sheet_name, index=False)
            
            self.logger.info(f"Fichier écrit avec succès: {len(df)} lignes")
            return True
            
        except Exception as e:
            self.logger.error(f"Erreur lors de l'écriture du fichier Excel: {e}")
            return False
    
    def write_formatted_excel(self, df: pd.DataFrame, file_path: str, 
                            sheet_name: str = "Updated_Data") -> bool:
        """
        Écrit un DataFrame avec formatage avancé.
        
        Args:
            df: DataFrame à écrire
            file_path: Chemin du fichier de sortie
            sheet_name: Nom de la feuille
            
        Returns:
            True si succès, False sinon
        """
        try:
            self.logger.info(f"Écriture du fichier Excel formaté: {file_path}")
            
            # Créer le répertoire si nécessaire
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Créer un nouveau workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            
            # Formater l'en-tête
            self._format_header(worksheet, len(df.columns))
            
            # Appliquer le formatage conditionnel
            if 'Action' in df.columns:
                self._apply_conditional_formatting(worksheet, df)
            
            # Ajouter des commentaires si présents
            if 'Notes' in df.columns:
                self._add_comments(worksheet, df)
            
            # Ajuster la largeur des colonnes
            self._adjust_column_widths(worksheet)
            
            # Sauvegarder le fichier
            workbook.save(file_path)
            
            self.logger.info(f"Fichier formaté écrit avec succès: {len(df)} lignes")
            return True
            
        except Exception as e:
            self.logger.error(f"Erreur lors de l'écriture du fichier formaté: {e}")
            return False
    
    def _format_header(self, worksheet, num_columns: int):
        """Formate l'en-tête du tableau."""
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _apply_conditional_formatting(self, worksheet, df: pd.DataFrame):
        """Applique le formatage conditionnel basé sur la colonne Action."""
        if 'Action' not in df.columns:
            return
        
        action_col_idx = df.columns.get_loc('Action') + 1
        
        for row_idx, action in enumerate(df['Action'], start=2):
            if pd.isna(action):
                continue
                
            action_str = str(action)
            
            # Déterminer la couleur basée sur l'action
            fill_color = None
            if action_str in ['Duplicate_Added', 'Unknown_Added']:
                fill_color = self.colors['duplicate']
            elif action_str == 'Updated':
                fill_color = self.colors['updated']
            elif action_str == 'Skipped':
                fill_color = self.colors['skipped']
            elif 'Error' in action_str:
                fill_color = self.colors['error']
            
            # Appliquer la couleur à toute la ligne
            if fill_color:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill_color
    
    def _add_comments(self, worksheet, df: pd.DataFrame):
        """Ajoute des commentaires aux cellules."""
        if 'Notes' not in df.columns:
            return
        
        notes_col_idx = df.columns.get_loc('Notes') + 1
        
        for row_idx, note in enumerate(df['Notes'], start=2):
            if pd.isna(note) or not str(note).strip():
                continue
            
            cell = worksheet.cell(row=row_idx, column=notes_col_idx)
            comment = Comment(str(note), "Component Processor")
            cell.comment = comment
    
    def _adjust_column_widths(self, worksheet):
        """Ajuste automatiquement la largeur des colonnes."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Largeur max de 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, workbook, summary_data: Dict[str, Any]):
        """
        Crée une feuille de résumé dans le workbook.
        
        Args:
            workbook: Workbook openpyxl
            summary_data: Données de résumé
        """
        try:
            # Créer une nouvelle feuille
            summary_sheet = workbook.create_sheet("Summary")
            
            # Ajouter le titre
            summary_sheet['A1'] = "RÉSUMÉ DU TRAITEMENT"
            summary_sheet['A1'].font = Font(size=16, bold=True)
            summary_sheet['A1'].alignment = Alignment(horizontal='center')
            
            # Fusionner les cellules pour le titre
            summary_sheet.merge_cells('A1:B1')
            
            # Ajouter les données de résumé
            row = 3
            for key, value in summary_data.items():
                summary_sheet[f'A{row}'] = key.replace('_', ' ').title()
                summary_sheet[f'B{row}'] = value
                summary_sheet[f'A{row}'].font = self.fonts['bold']
                row += 1
            
            # Ajuster les largeurs
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 15
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la création de la feuille de résumé: {e}")
    
    def export_multiple_sheets(self, data_dict: Dict[str, pd.DataFrame], 
                              file_path: str, summary_data: Optional[Dict] = None) -> bool:
        """
        Exporte plusieurs DataFrames dans un fichier Excel multi-feuilles.
        
        Args:
            data_dict: Dictionnaire {nom_feuille: DataFrame}
            file_path: Chemin du fichier de sortie
            summary_data: Données de résumé optionnelles
            
        Returns:
            True si succès, False sinon
        """
        try:
            self.logger.info(f"Export multi-feuilles vers: {file_path}")
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in data_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Ajouter une feuille de résumé si fournie
                if summary_data:
                    self.create_summary_sheet(writer.book, summary_data)
            
            self.logger.info("Export multi-feuilles terminé avec succès")
            return True
            
        except Exception as e:
            self.logger.error(f"Erreur lors de l'export multi-feuilles: {e}")
            return False
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Retourne la liste des noms de feuilles d'un fichier Excel.
        
        Args:
            file_path: Chemin du fichier Excel
            
        Returns:
            Liste des noms de feuilles
        """
        try:
            excel_file = pd.ExcelFile(file_path)
            return excel_file.sheet_names
        except Exception as e:
            self.logger.error(f"Erreur lors de la lecture des noms de feuilles: {e}")
            return []
